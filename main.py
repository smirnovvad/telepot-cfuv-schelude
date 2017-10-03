import sys
import asyncio
import telepot.aio
from telepot import glance
from telepot.aio.loop import MessageLoop
from telepot.namedtuple import InlineKeyboardMarkup, InlineKeyboardButton, ReplyKeyboardMarkup
from telepot.aio.delegate import (
    per_chat_id, per_callback_query_origin, create_open, pave_event_space)
import datetime
import config
import img
from openpyxl import load_workbook


def render(group, day):
    try:
        wb = load_workbook('xlsx/%s.xlsx' % (day))
        ws = wb['Расписание (%s курс)' % (group[0])]

        lessons = group + '\n'
        aud = '\n'

        for x in range(3, 10):
            if ws.cell(row=x, column=config.courses[group[0]][group]).value is None:
                lessons += '\n'
            else:
                lessons += ws.cell(row=x, column=config.courses[group[0]][group]).value.replace('\n', ' / ') + '\n'
            if ws.cell(row=x, column=config.courses[group[0]][group] + 1).value is None:
                aud += '\n'
            else:
                aud += ws.cell(row=x, column=config.courses[group[0]][group] + 1).value.replace('\n', ' / ') + '\n'
        img.render([lessons, aud], (85, 85, 85), group + day)
    except Exception as e:
        print(e)


class ScheluderStarter(telepot.aio.helper.ChatHandler):

    def __init__(self, *args, **kwargs):
        super(ScheluderStarter, self).__init__(*args, **kwargs)

    async def on_chat_message(self, msg):
        content_type, chat_type, chat_id = glance(msg)

        if msg['from']['id'] in config.admins and content_type == 'document' and 'xlsx' in msg['document']['file_name']:
            if any(x.lower() in msg['document']['file_name'] for x in config.days.values()):
                print(msg['document']['file_name'])
                for day in config.days.values():
                    if day.lower() in msg['document']['file_name']:
                        await bot.download_file(msg['document']['file_id'], 'xlsx/' + day[:1].upper() + day[1:] + '.xlsx')
                        await self.sender.sendMessage(
                            'Файл сохранен как ' + day,
                        )
        if content_type == 'text':
            print('Chat:', content_type, chat_type, msg['text'], msg['from'], datetime.datetime.fromtimestamp(msg['date']).strftime('%Y-%m-%d %H:%M:%S'))
            await self.sender.sendMessage(
                'Привет, выбери курс',
                reply_markup=InlineKeyboardMarkup(
                    inline_keyboard=[[
                        InlineKeyboardButton(text='1 Курс', callback_data='1'),
                        InlineKeyboardButton(text='2 Курс', callback_data='2'),
                        InlineKeyboardButton(text='3 Курс', callback_data='3'),
                        InlineKeyboardButton(text='4 Курс', callback_data='4'),
                    ]]
                )
            )
        self.close()


class Scheluder(telepot.aio.helper.CallbackQueryOriginHandler):

    def __init__(self, *args, **kwargs):
        super(Scheluder, self).__init__(*args, **kwargs)
        self._course = 0
        self._group = ''
        self._day = ''

    async def _show_groups(self, groups):
        await self.editor.editMessageText('Выбери группу',
                                          reply_markup=InlineKeyboardMarkup(
                                              inline_keyboard=list(map(lambda c: [InlineKeyboardButton(
                                                  text=str(c), callback_data=str(c))], list(groups.keys())))

                                          )
                                          )

    async def _show_days(self, days):
        await self.editor.editMessageText('Выбери день недели',
                                          reply_markup=InlineKeyboardMarkup(
                                              inline_keyboard=list(map(lambda c: [InlineKeyboardButton(
                                                  text=str(c), callback_data=str(c))], list(days.values())))

                                          )
                                          )

    async def on_callback_query(self, msg):
        query_id, from_id, query_data = glance(msg, flavor='callback_query')
        self.sender = telepot.aio.helper.Sender(self.bot, from_id)
        if query_data in config.courses:
            self._course = query_data
            await self._show_groups(config.courses[self._course])

        elif query_data in config.days.values():
            self._day = query_data
            render(self._group, self._day)
            print(self._group, self._day, datetime.datetime.today())
            await self.sender.sendPhoto(open('img/%s.png' % (self._group + self._day), 'rb'))
            await self.sender.sendMessage(
                'Привет, выбери курс',
                reply_markup=InlineKeyboardMarkup(
                    inline_keyboard=[[
                        InlineKeyboardButton(text='1 Курс', callback_data='1'),
                        InlineKeyboardButton(text='2 Курс', callback_data='2'),
                        InlineKeyboardButton(text='3 Курс', callback_data='3'),
                        InlineKeyboardButton(text='4 Курс', callback_data='4'),
                    ]]
                )
            )

        elif query_data in config.courses[self._course]:
            print(query_data)
            self._group = query_data
            await self._show_days(config.days)

    async def on__idle(self, event):
        await asyncio.sleep(5)
        await self.editor.deleteMessage()

        self.close()


def chunks(lst, chunk_count):
    chunk_size = len(lst) // chunk_count
    return [lst[i:i + chunk_size] for i in range(0, len(lst), chunk_size)]


TOKEN = sys.argv[1]  # get token from command-line

bot = telepot.aio.DelegatorBot(TOKEN, [
    pave_event_space()(
        per_chat_id(), create_open, ScheluderStarter, timeout=3),
    pave_event_space()(
        per_callback_query_origin(), create_open, Scheluder, timeout=60),
])

loop = asyncio.get_event_loop()

loop.create_task(MessageLoop(bot).run_forever())
print('Listening ...')

loop.run_forever()
